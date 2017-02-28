#-*- coding: utf-8 -*-

import xlrd
import openpyxl
from openpyxl import Workbook


def excelRead(filepath, sheetname):
    workbook = xlrd.open_workbook(filepath)
    worksheet = workbook.sheet_by_name(sheetname)
    
    num_rows = worksheet.nrows -1
    curr_row = -1
    result = []
    
    while curr_row < num_rows:
        curr_row += 1
        row = worksheet.row(curr_row)
        result.append(row)
    
    return result

def excelWriteOnExistingFile(filepath, sheetname, startCol, insert):
    # insert - double list [ [], [] , [], ...]
    
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_name(sheetname)
    
    workbook = openpyxl.load_workbook(filepath)
    worksheet = workbook.active
    
    
    rowLen = ws.nrows
    
    strtColNum = ord(startCol)

    # row
    indRow = 0
    for lst in insert:
    # cell
        indCol = 0
        for attr in lst:
            col = strtColNum+indCol
            row = rowLen+1+indRow
            
            octave = 0
            if col > 90:
                octave += 1
                col -= 26
            octaveChr = ''
            if octave > 0:
                octaveChr = chr(64+octave)
            
            try:
                worksheet[octaveChr+chr(col)+str(row)] = attr
            except(TypeError):
                print ('Type Error - '+str(indCol))
            
            indCol+=1
        indRow+=1
    workbook.save(filepath)
    print('saved successfully in existing file!') 
    
    
    
def excelWriteNewFile(filepath, sheetname, insertList):
    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    
    i1 = 0
    i2 = ord('A')
    while i1<len(insertList[0]):
        j=0
        k=j+1
        while j<len(insertList):
            #print(type(insertList[j][i1]))
            if str(type(insertList[j][i1])) == "<class 'xlrd.sheet.Cell'>":
                ws[chr(i2)+str(k)] = insertList[j][i1].value
            else:
                ws[chr(i2)+str(k)] = insertList[j][i1].encode('utf-8')
            j+=1
            k+=1
        i1+=1
        i2+=1
        
    wb.save(filepath)
    print('saved successfully in a new file!')
    